from table import Table, Seat
import numpy as np
import openpyxl

class Openspace:

    def __init__(self, number_of_tables, seats_per_table):
        self.number_of_tables = number_of_tables
        self.tables = [Table(seats_per_table) for _ in range(number_of_tables)]

    def organize(self, names: str):
        """
        Randomly assign people to Seat objects in the different Table objects.
        """
        np.random.shuffle(names)
        name_index = 0
        total_seats = sum(len(table.seats) for table in self.tables)

        if len(names) > total_seats:
            raise ValueError("There are more names than seats available.")

        for table in self.tables:
            for seat in table.seats:
                if name_index < len(names):
                    seat.occupant = names[name_index]
                    name_index += 1
                else:
                    seat.occupant = None  # If there are more seats than names

    def display(self):
        """
        Display the different tables and their occupants in a nice and readable way.
        """
        for idx, table in enumerate(self.tables):
            print(f"Table {idx + 1} has:")
            for seat in table.seats:
                occupant = seat.occupant if seat.occupant else "Empty"
                print(f"  - {occupant}")

    def store(self, filename='e.xlsx'):
        """
        Store the repartition in an Excel file.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        for idx, table in enumerate(self.tables):
            sheet.cell(row=row, column=1, value=f"Table {idx + 1}")
            row += 1
            for seat in table.seats:
                occupant = seat.occupant if seat.occupant else "Empty"
                sheet.cell(row=row, column=1, value=occupant)
                row += 1
            row += 1  # Add an empty row between tables

        workbook.save(filename)
        return f"Repartition saved successfully in {filename}."

a = Openspace(3,3)
print(a.organize(["A","B","C","D","E","F",'R','Y','O']))
print(a.display())
print(a.store())
